from splinter import Browser
import time,sys
import openpyxl as xl

def overFillById(browser,id,fillString):
	if browser.is_element_present_by_id(id,wait_time=2):
		browser.find_by_id(id).fill(fillString)
	return;

def overClickById(browser,id):
	if browser.is_element_present_by_id(id,wait_time=2):
		browser.find_by_id(id).click() 
	return;

def overClickByText(browser,text):
	if browser.is_element_present_by_text(text,wait_time=2):
		browser.find_by_text(text).click() 
	return;
	
def getStrTiming(hStr):
	vTiming = ['', '']
	if float(hStr) > 9:
		return vTiming
		
	if float(hStr) % 0.5 > 0:
		return vTiming
		
	integer = int(hStr)
	decimal = float(hStr) % 1
	
	if integer > 0 or decimal > 0:
		vTiming[0] = '00:00'
		vTiming[1] = '0' + str(integer) + ':' + str(int(decimal * 60)).zfill(2)
		
	return vTiming;

def insertNavRecord(browser,date,codiceCommessa,faseCommessa,hBase,hStr ):
	vTiming = getStrTiming(hStr)
	time.sleep(1)
	overClickById(browser,'listaRigheOdT_btnNew')
	time.sleep(1)
	overClickById(browser,'btnRicercaCommessa')
	time.sleep(2)
	overFillById(browser,'searchCodCommessa',codiceCommessa)
	overFillById(browser,'searchFase',faseCommessa)
	browser.find_by_id('listaCommesse_btnFiltra').click()
	time.sleep(1)
	overClickByText(browser,codiceCommessa)
	time.sleep(2)
	overFillById(browser,'Data',date)
	overClickById(browser,'Note')
	overFillById(browser,'hord',hBase)
	time.sleep(1)
	overFillById(browser,'tstraini',vTiming[0])
	time.sleep(1)
	overFillById(browser,'tstrafin',vTiming[1])
	time.sleep(1)	
	overClickById(browser,'btnSalva')
	time.sleep(2)
	return;

def navisionLogin(browser):
	browser.visit('https://navisionweb.lutech.it')
	browser.find_by_id('UserName').fill(sys.argv[1])
	browser.find_by_id('Password').fill(sys.argv[2])
	browser.find_by_id('btnLogin').click()
	return;

def newReport(browser):
	browser.find_by_id('menu1').click()
	time.sleep(1)
	if browser.is_element_present_by_id('listaOdT_btnNew',wait_time=2):
		browser.find_by_id('listaOdT_btnNew').click()
	time.sleep(1)
	return;  

def openBozzaReport(browser):
	browser.find_by_id('menu1').click()
	browser.find_by_text('BOZZA').click()
	time.sleep(1)
	if browser.is_element_present_by_id('listaOdT_btnEdit',wait_time=2):
		browser.find_by_id('listaOdT_btnEdit').click()
	time.sleep(1)
	return;
	
if not sys.argv[1] or not sys.argv[2]:
	exit()
	
browser = Browser('chrome')
navisionLogin(browser)
newReport(browser)

wb = xl.load_workbook(filename = 'Navision.xlsm')
ws = wb.active
i = 4
auxDate = ws.cell(row=2, column=i).value

while auxDate:
	print('Loading ' + auxDate)
	j = 3
	auxProgram = ws.cell(row=j, column=2).value
	auxPhase = ws.cell(row=j, column=3).value
	
	while auxProgram:
		#print(auxProgram + ':' + auxPhase)
		auxHours = ws.cell(row=j, column=i).value
		if auxHours:
			#print(auxHours)
			auxHours = str(auxHours).split('+')		
			if len(auxHours) == 1:
				auxHours.append('0')
				
			if not auxHours[1]:
				auxHours[1] = '0'
				
			if float(auxHours[0]) + float(auxHours[1]) > 0:
				#print(auxHours[0] + ' ' + auxHours[1])
				insertNavRecord(browser,auxDate,auxProgram,auxPhase,auxHours[0],auxHours[1])
		j += 1
		auxProgram = ws.cell(row=j, column=2).value
		auxPhase = ws.cell(row=j, column=3).value
		
	i += 1
	auxDate = ws.cell(row=2, column=i).value

browser.quit()
print('Load complete')
